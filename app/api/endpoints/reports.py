from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, case, or_
from typing import List, Optional
from uuid import UUID
from datetime import date, timedelta
from io import BytesIO
from decimal import Decimal
from app.db.session import get_db
from app.models.user import User
from app.models.task_entry import TaskEntry, TaskEntryStatus, TaskSubEntry
from app.models.leave_request import LeaveRequest, LeaveStatus
from app.models.client import Client
from app.models.task_master import TaskMaster
from app.models.holiday import Holiday
from app.models.working_saturday import WorkingSaturday
from app.schemas import TimesheetReportItem, AttendanceReportItem, LeaveReportItem, ProductionReportItem
from app.api.dependencies import require_admin, get_current_user
from app.core.date_filters import get_date_range

router = APIRouter(prefix="/admin/reports", tags=["Admin - Reports"])


@router.get("/timesheet", response_model=List[ProductionReportItem])
def get_timesheet_report(
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    date_range: Optional[str] = Query(None),
    user_ids: Optional[str] = Query(None),  # Comma-separated UUIDs
    client_ids: Optional[str] = Query(None),  # Comma-separated UUIDs
    task_master_ids: Optional[str] = Query(None),  # Comma-separated UUIDs
    is_profitable: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin)
):
    """Get client-wise production report (admin only) grouped by client, employee, task."""
    # Handle date_range filter
    if not from_date or not to_date:
        if date_range and date_range != "custom":
            from_date, to_date = get_date_range(date_range)
        else:
            from_date, to_date = get_date_range("this_month")

    query = db.query(
        func.coalesce(Client.name, 'No Client').label('client_name'),
        User.name.label('employee_name'),
        TaskMaster.name.label('task_name'),
        func.sum(TaskSubEntry.production).label('total_production'),
        func.sum(TaskSubEntry.hours).label('total_hours'),
    ).join(
        TaskEntry, TaskSubEntry.task_entry_id == TaskEntry.id
    ).outerjoin(
        Client, TaskSubEntry.client_id == Client.id
    ).join(
        TaskMaster, TaskSubEntry.task_master_id == TaskMaster.id
    ).join(
        User, TaskEntry.user_id == User.id
    ).filter(
        TaskEntry.work_date >= from_date,
        TaskEntry.work_date <= to_date,
        # TaskEntry.status == TaskEntryStatus.APPROVED,
        User.role == 'EMPLOYEE',
    )

    # Multi-select filters
    if user_ids:
        user_id_list = [uid.strip() for uid in user_ids.split(',') if uid.strip()]
        if user_id_list:
            query = query.filter(TaskEntry.user_id.in_(user_id_list))

    if client_ids:
        client_id_list = [cid.strip() for cid in client_ids.split(',') if cid.strip()]
        if client_id_list:
            query = query.filter(TaskSubEntry.client_id.in_(client_id_list))

    if task_master_ids:
        task_master_id_list = [tmid.strip() for tmid in task_master_ids.split(',') if tmid.strip()]
        if task_master_id_list:
            query = query.filter(TaskSubEntry.task_master_id.in_(task_master_id_list))

    if is_profitable is not None:
        query = query.filter(TaskMaster.is_profitable == is_profitable)

    query = query.group_by(
        Client.name, User.name, TaskMaster.name
    ).order_by(Client.name, User.name, TaskMaster.name)

    results = query.all()

    return [
        ProductionReportItem(
            client_name=row.client_name,
            employee_name=row.employee_name,
            task_name=row.task_name,
            production=float(row.total_production or 0),
            hours=float(row.total_hours or 0),
            efficiency=round(float(row.total_production or 0) / float(row.total_hours) if row.total_hours else 0, 2),
        )
        for row in results
    ]


@router.get("/attendance", response_model=List[AttendanceReportItem])
def get_attendance_report(
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    date_range: Optional[str] = Query(None),
    user_ids: Optional[str] = Query(None),
    is_profitable: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin)
):

    if date_range:
        from_date, to_date = get_date_range(date_range)
    elif not from_date or not to_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Either provide from_date and to_date, or date_range"
        )

    # ── Load Users ────────────────────────────────────────────────
    user_query = db.query(User).filter(
        User.is_active == True,
        User.role.in_(["EMPLOYEE", "SUPERVISOR"])
    )

    if user_ids:
        uid_list = [u.strip() for u in user_ids.split(",") if u.strip()]
        if uid_list:
            user_query = user_query.filter(User.id.in_(uid_list))

    users = user_query.order_by(User.name).all()
    user_id_strs = [str(u.id) for u in users]

    # ── Holidays ─────────────────────────────────────────────────
    holiday_map = {
        row.holiday_date: row.name
        for row in db.query(Holiday.holiday_date, Holiday.name).filter(
            Holiday.holiday_date >= from_date,
            Holiday.holiday_date <= to_date
        ).all()
    }

    # ── Working Saturdays ───────────────────────────────────────
    working_saturday_set = {
        row.work_date
        for row in db.query(WorkingSaturday.work_date).filter(
            WorkingSaturday.work_date >= from_date,
            WorkingSaturday.work_date <= to_date
        ).all()
    }

    # ── Leave Task Master (Exact Match) ─────────────────────────
    leave_tm_ids = {
        row.id for row in db.query(TaskMaster.id).filter(
            func.lower(TaskMaster.name) == "leave"
        ).all()
    }

    # ── Leave Hours Aggregation ─────────────────────────────────
    from collections import defaultdict
    leave_hours_map = defaultdict(float)

    leave_q = db.query(
        TaskEntry.user_id,
        TaskEntry.work_date,
        func.sum(TaskSubEntry.hours).label("total_leave_hours")
    ).join(
        TaskSubEntry, TaskEntry.id == TaskSubEntry.task_entry_id
    ).filter(
        TaskEntry.work_date >= from_date,
        TaskEntry.work_date <= to_date,
        TaskEntry.status == TaskEntryStatus.APPROVED,
        TaskEntry.user_id.in_(user_id_strs),
        TaskSubEntry.task_master_id.in_(leave_tm_ids),
    ).group_by(
        TaskEntry.user_id,
        TaskEntry.work_date
    )

    for row in leave_q.all():
        leave_hours_map[(str(row.user_id), row.work_date)] = float(row.total_leave_hours or 0)

    # ── Entry Status Map ────────────────────────────────────────
    STATUS_PRIORITY = {
        TaskEntryStatus.APPROVED: 4,
        TaskEntryStatus.PENDING: 3,
        TaskEntryStatus.DRAFT: 2,
        TaskEntryStatus.REJECTED: 1,
    }

    entry_map = {}

    for er in db.query(TaskEntry.user_id, TaskEntry.work_date, TaskEntry.status).filter(
        TaskEntry.work_date >= from_date,
        TaskEntry.work_date <= to_date,
        TaskEntry.user_id.in_(user_id_strs)
    ).all():

        key = (str(er.user_id), er.work_date)
        if key not in entry_map or STATUS_PRIORITY.get(er.status, 0) > STATUS_PRIORITY.get(entry_map[key], 0):
            entry_map[key] = er.status

    # ── Production Map ──────────────────────────────────────────
    prod_map = defaultdict(lambda: {"production": 0.0, "clients": []})

    prod_q = db.query(
        TaskEntry.user_id,
        TaskEntry.work_date,
        func.coalesce(Client.name, "No Client").label("client_name"),
        func.sum(TaskSubEntry.production).label("total_production"),
    ).join(
        TaskSubEntry, TaskEntry.id == TaskSubEntry.task_entry_id
    ).outerjoin(
        Client, TaskSubEntry.client_id == Client.id
    ).filter(
        TaskEntry.work_date >= from_date,
        TaskEntry.work_date <= to_date,
        TaskEntry.status == TaskEntryStatus.APPROVED,
        TaskEntry.user_id.in_(user_id_strs),
    )

    if is_profitable is not None:
        prod_q = prod_q.join(
            TaskMaster, TaskSubEntry.task_master_id == TaskMaster.id
        ).filter(TaskMaster.is_profitable == is_profitable)

    prod_q = prod_q.group_by(TaskEntry.user_id, TaskEntry.work_date, Client.name)

    for pr in prod_q.all():
        key = (str(pr.user_id), pr.work_date)
        prod_map[key]["production"] += float(pr.total_production or 0)
        if pr.client_name and pr.client_name not in prod_map[key]["clients"]:
            prod_map[key]["clients"].append(pr.client_name)

    # ── Build Report ─────────────────────────────────────────────
    report_items = []
    current_date = from_date
    while current_date <= to_date:

        weekday = current_date.weekday()
        is_sunday = weekday == 6
        is_saturday = weekday == 5
        is_working_saturday = current_date in working_saturday_set
        in_holiday_map = current_date in holiday_map

        # Check if current date is a non-working day (weekend/holiday)
        is_non_working_day = is_sunday or in_holiday_map or (is_saturday and not is_working_saturday)

        for user in users:

            uid = str(user.id)
            entry_key = (uid, current_date)

            pd = prod_map.get(entry_key)
            leave_hours = leave_hours_map.get(entry_key, 0)

            is_full_day_leave = False
            is_half_day_leave = False
            is_short_leave = False
            has_production = pd and pd["production"] > 0

            if user.role == "SUPERVISOR":
                has_production = pd and pd["production"] > 0
                attendance_status = "PRESENT"
                production = round(pd["production"], 2) if pd else 0.0
                client_names_str = ", ".join(pd["clients"]) if pd and pd["clients"] else None

                # Mark SUPERVISOR as OVERTIME if working on non-working day (weekend/holiday)
                if is_non_working_day:
                    attendance_status = "OVERTIME"

            elif entry_key in entry_map and entry_map[entry_key] == TaskEntryStatus.APPROVED:

                # Categorize Leave
                if leave_hours >= 8 or leave_hours > 4:
                    is_full_day_leave = True
                elif leave_hours > 2:
                    is_half_day_leave = True
                elif leave_hours > 0:
                    is_short_leave = True

                if has_production:
                    attendance_status = "PRESENT"
                    production = round(pd["production"], 2)
                    client_names_str = ", ".join(pd["clients"]) if pd["clients"] else None
                elif leave_hours > 0:
                    attendance_status = "LEAVE"
                    production = None
                    client_names_str = None
                else:
                    attendance_status = "ABSENT"
                    production = None
                    client_names_str = None

                # Mark as OVERTIME if working on non-working day (weekend/holiday)
                if is_non_working_day and has_production:
                    attendance_status = "OVERTIME"

            else:
                attendance_status = "ABSENT"
                production = None
                client_names_str = None

            # Skip ABSENT records on non-working days (treat as holidays)
            if is_non_working_day and attendance_status == "ABSENT":
                continue

            report_items.append(AttendanceReportItem(
                date=current_date,
                employee_name=user.name,
                employee_email=user.email,
                attendance_status=attendance_status,
                is_holiday=False,
                production=production,
                client_names=client_names_str,
                is_full_day_leave=is_full_day_leave,
                is_half_day_leave=is_half_day_leave,
                is_short_leave=is_short_leave,
                leave_hours=leave_hours if leave_hours > 0 else None
            ))

        current_date += timedelta(days=1)

    return report_items

@router.get("/leave", response_model=List[LeaveReportItem])
def get_leave_report(
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    date_range: Optional[str] = Query(None),
    user_id: Optional[UUID] = Query(None),
    is_profitable: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin)
):
    """Get leave report (admin only)."""
    # Handle date_range filter
    if date_range:
        from_date, to_date = get_date_range(date_range)
    elif not from_date or not to_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Either provide from_date and to_date, or date_range"
        )
    query = db.query(
        LeaveRequest.from_date,
        LeaveRequest.to_date,
        User.name,
        User.email,
        LeaveRequest.reason,
        LeaveRequest.status,
        LeaveRequest.admin_comment
    ).join(User, LeaveRequest.user_id == User.id)
    
    # Apply filters - get leaves that overlap with the date range
    query = query.filter(
        LeaveRequest.from_date <= to_date,
        LeaveRequest.to_date >= from_date
    )
    
    # Filter out admin users - only show employee records
    query = query.filter(User.role == 'EMPLOYEE')
    
    if user_id:
        query = query.filter(LeaveRequest.user_id == user_id)
    # Note: is_profitable filter doesn't apply to leave requests
    
    results = query.order_by(LeaveRequest.from_date.desc()).all()
    
    return [
        LeaveReportItem(
            from_date=row[0],
            to_date=row[1],
            employee_name=row[2],
            employee_email=row[3],
            reason=row[4],
            status=row[5],
            admin_comment=row[6]
        )
        for row in results
    ]


@router.get("/export/attendance-excel")
async def export_attendance_to_excel(
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    date_range: Optional[str] = Query("current_week"),
    user_ids: Optional[str] = Query(None),
    is_profitable: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl library is not installed"
        )

    if date_range and date_range != "custom":
        from_date, to_date = get_date_range(date_range)
    elif not from_date or not to_date:
        from_date, to_date = get_date_range("current_week")

    # ───────── Users ─────────
    user_query = db.query(User).filter(
        User.is_active == True,
        User.role.in_(["EMPLOYEE", "SUPERVISOR"])
    )

    if user_ids:
        uid_list = [u.strip() for u in user_ids.split(",") if u.strip()]
        if uid_list:
            user_query = user_query.filter(User.id.in_(uid_list))

    users = user_query.order_by(User.name).all()
    user_id_strs = [str(u.id) for u in users]

    # ───────── Holidays / Saturdays ─────────
    holiday_map = {
        row.holiday_date: row.name
        for row in db.query(Holiday.holiday_date, Holiday.name).filter(
            Holiday.holiday_date >= from_date,
            Holiday.holiday_date <= to_date
        ).all()
    }

    working_saturday_set = {
        row.work_date
        for row in db.query(WorkingSaturday.work_date).filter(
            WorkingSaturday.work_date >= from_date,
            WorkingSaturday.work_date <= to_date
        ).all()
    }

    # ───────── Leave Task Master ─────────
    leave_tm_ids = {
        row.id for row in db.query(TaskMaster.id).filter(
            func.lower(TaskMaster.name) == "leave"
        ).all()
    }

    from collections import defaultdict

    # ───────── Leave Hours Map ─────────
    leave_hours_map = defaultdict(float)

    leave_q = db.query(
        TaskEntry.user_id,
        TaskEntry.work_date,
        func.sum(TaskSubEntry.hours).label("total_leave_hours")
    ).join(
        TaskSubEntry, TaskEntry.id == TaskSubEntry.task_entry_id
    ).filter(
        TaskEntry.work_date >= from_date,
        TaskEntry.work_date <= to_date,
        TaskEntry.status == TaskEntryStatus.APPROVED,
        TaskEntry.user_id.in_(user_id_strs),
        TaskSubEntry.task_master_id.in_(leave_tm_ids),
    ).group_by(
        TaskEntry.user_id,
        TaskEntry.work_date
    )

    for row in leave_q.all():
        leave_hours_map[(str(row.user_id), row.work_date)] = float(row.total_leave_hours or 0)

    # ───────── Entry Map ─────────
    STATUS_PRIORITY = {
        TaskEntryStatus.APPROVED: 4,
        TaskEntryStatus.PENDING: 3,
        TaskEntryStatus.DRAFT: 2,
        TaskEntryStatus.REJECTED: 1,
    }

    entry_map = {}

    for er in db.query(TaskEntry.user_id, TaskEntry.work_date, TaskEntry.status).filter(
        TaskEntry.work_date >= from_date,
        TaskEntry.work_date <= to_date,
        TaskEntry.user_id.in_(user_id_strs)
    ).all():

        key = (str(er.user_id), er.work_date)
        if key not in entry_map or STATUS_PRIORITY.get(er.status, 0) > STATUS_PRIORITY.get(entry_map[key], 0):
            entry_map[key] = er.status

    # ───────── Production Map ─────────
    prod_map = defaultdict(lambda: {"production": 0.0, "clients": []})

    prod_q = db.query(
        TaskEntry.user_id,
        TaskEntry.work_date,
        func.coalesce(Client.name, "No Client").label("client_name"),
        func.sum(TaskSubEntry.production).label("total_production"),
    ).join(
        TaskSubEntry, TaskEntry.id == TaskSubEntry.task_entry_id
    ).outerjoin(
        Client, TaskSubEntry.client_id == Client.id
    ).filter(
        TaskEntry.work_date >= from_date,
        TaskEntry.work_date <= to_date,
        TaskEntry.status == TaskEntryStatus.APPROVED,
        TaskEntry.user_id.in_(user_id_strs),
    )

    if is_profitable is not None:
        prod_q = prod_q.join(
            TaskMaster, TaskSubEntry.task_master_id == TaskMaster.id
        ).filter(TaskMaster.is_profitable == is_profitable)

    prod_q = prod_q.group_by(TaskEntry.user_id, TaskEntry.work_date, Client.name)

    for pr in prod_q.all():
        key = (str(pr.user_id), pr.work_date)
        prod_map[key]["production"] += float(pr.total_production or 0)
        if pr.client_name and pr.client_name not in prod_map[key]["clients"]:
            prod_map[key]["clients"].append(pr.client_name)

    # ───────── Build Rows ─────────
    rows = []
    current_date = from_date

    while current_date <= to_date:

        weekday = current_date.weekday()
        is_sunday = weekday == 6
        is_saturday = weekday == 5
        is_working_saturday = current_date in working_saturday_set
        in_holiday_map = current_date in holiday_map

        if is_sunday or in_holiday_map or (is_saturday and not is_working_saturday):
            current_date += timedelta(days=1)
            continue

        for user in users:

            uid = str(user.id)
            entry_key = (uid, current_date)

            pd = prod_map.get(entry_key)
            leave_hours = leave_hours_map.get(entry_key, 0)

            status = "ABSENT"
            production = None
            clients = None

            if user.role == "SUPERVISOR":
                status = "PRESENT"
                production = round(pd["production"], 2) if pd else 0.0
                clients = ", ".join(pd["clients"]) if pd and pd["clients"] else None

            elif entry_key in entry_map and entry_map[entry_key] == TaskEntryStatus.APPROVED:

                has_production = pd and pd["production"] > 0

                leave_label = None
                if leave_hours >= 8 or leave_hours > 4:
                    leave_label = "FULL DAY LEAVE"
                elif leave_hours > 2:
                    leave_label = "HALF DAY LEAVE"
                elif leave_hours > 0:
                    leave_label = "SHORT LEAVE"

                if has_production:
                    status = "PRESENT"
                    production = round(pd["production"], 2)
                    clients = ", ".join(pd["clients"]) if pd["clients"] else None
                    if leave_label:
                        status = leave_label
                elif leave_label:
                    status = leave_label
                else:
                    status = "ABSENT"

            rows.append({
                "date": str(current_date),
                "employee": user.name,
                "status": status,
                "production": production,
                "clients": clients
            })

        current_date += timedelta(days=1)

    # ───────── Excel Generation ─────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ["Date", "Employee", "Status", "Production", "Client(s)"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    row_num = 2
    for r in rows:
        values = [r["date"], r["employee"], r["status"],
                  r["production"] if r["production"] else "",
                  r["clients"] or ""]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = border
        row_num += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 28

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    filename = f"attendance_report_{from_date}_{to_date}.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/export/excel")
async def export_to_excel(
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    date_range: Optional[str] = Query("this_month"),
    client_ids: Optional[str] = Query(None),  # Comma-separated UUIDs
    user_ids: Optional[str] = Query(None),  # Comma-separated UUIDs
    task_master_ids: Optional[str] = Query(None),  # Comma-separated UUIDs
    is_profitable: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """Export client-wise production report to Excel (admin only)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl library is not installed"
        )

    # Handle date_range filter
    if not from_date or not to_date:
        if date_range and date_range != "custom":
            from_date, to_date = get_date_range(date_range)
        else:
            from_date, to_date = get_date_range("this_month")

    # Query sub-entry level data grouped by client + employee + task
    query = db.query(
        func.coalesce(Client.name, 'No Client').label('client_name'),
        User.name.label('employee_name'),
        TaskMaster.name.label('task_name'),
        func.sum(TaskSubEntry.production).label('total_production'),
        func.sum(TaskSubEntry.hours).label('total_hours'),
    ).join(
        TaskEntry, TaskSubEntry.task_entry_id == TaskEntry.id
    ).outerjoin(
        Client, TaskSubEntry.client_id == Client.id
    ).join(
        TaskMaster, TaskSubEntry.task_master_id == TaskMaster.id
    ).join(
        User, TaskEntry.user_id == User.id
    ).filter(
        TaskEntry.work_date >= from_date,
        TaskEntry.work_date <= to_date,
        TaskEntry.status == TaskEntryStatus.APPROVED,
        User.role == 'EMPLOYEE',
    )

    if client_ids:
        client_id_list = [cid.strip() for cid in client_ids.split(',') if cid.strip()]
        if client_id_list:
            query = query.filter(TaskSubEntry.client_id.in_(client_id_list))

    if user_ids:
        user_id_list = [uid.strip() for uid in user_ids.split(',') if uid.strip()]
        if user_id_list:
            query = query.filter(TaskEntry.user_id.in_(user_id_list))

    if task_master_ids:
        task_master_id_list = [tmid.strip() for tmid in task_master_ids.split(',') if tmid.strip()]
        if task_master_id_list:
            query = query.filter(TaskSubEntry.task_master_id.in_(task_master_id_list))

    if is_profitable is not None:
        query = query.filter(TaskMaster.is_profitable == is_profitable)

    query = query.group_by(
        Client.name, User.name, TaskMaster.name
    ).order_by(Client.name, User.name, TaskMaster.name)

    results = query.all()

    # Build flat row list
    rows = []
    for row in results:
        prod = float(row.total_production or 0)
        hrs = float(row.total_hours or 0)
        efficiency = round(prod / hrs, 2) if hrs > 0 else 0
        rows.append({
            'client_name': row.client_name,
            'employee_name': row.employee_name,
            'task_name': row.task_name,
            'production': prod,
            'hours': hrs,
            'efficiency': efficiency,
        })

    # ── Excel workbook ──────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet Report"

    # Styles
    title_fill   = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_font   = Font(color="FFFFFF", bold=True, size=14)
    header_fill  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font  = Font(color="FFFFFF", bold=True)
    total_fill   = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    total_font   = Font(bold=True, size=12)
    center       = Alignment(horizontal="center", vertical="center")
    border       = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )

    NUM_COLS = 6
    row_num = 1

    # Title row
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=NUM_COLS)
    tc = ws.cell(row=row_num, column=1, value=f"Timesheet Report: {from_date} to {to_date}")
    tc.fill, tc.font, tc.alignment = title_fill, title_font, center
    row_num += 2

    # Column header row
    headers = ["Client", "Employee", "Task Name", "Production", "Hours", "Efficiency"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=header)
        cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, center, border
    row_num += 1

    if not rows:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=NUM_COLS)
        nd = ws.cell(row=row_num, column=1, value="No data found for the selected date range and filters")
        nd.alignment = center
        nd.font = Font(italic=True, size=12)
    else:
        grand_production = 0.0
        grand_hours = 0.0

        for r in rows:
            ws.cell(row=row_num, column=1, value=r['client_name']).border = border
            ws.cell(row=row_num, column=2, value=r['employee_name']).border = border
            ws.cell(row=row_num, column=3, value=r['task_name']).border = border
            ws.cell(row=row_num, column=4, value=r['production']).border = border
            ws.cell(row=row_num, column=5, value=r['hours']).border = border
            ws.cell(row=row_num, column=6, value=r['efficiency']).border = border
            row_num += 1
            grand_production += r['production']
            grand_hours += r['hours']

        # Grand total row
        row_num += 1
        grand_efficiency = round(grand_production / grand_hours, 2) if grand_hours > 0 else 0

        labels = ["TOTAL", "", "", round(grand_production, 2), round(grand_hours, 2), grand_efficiency]
        for col, val in enumerate(labels, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill, cell.font, cell.alignment, cell.border = total_fill, total_font, center, border

    # Column widths
    col_widths = [25, 20, 25, 14, 12, 14]
    for col, width in zip("ABCDEF", col_widths):
        ws.column_dimensions[col].width = width

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    filename = f"timesheet_report_{from_date}_{to_date}.xlsx"
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/leave-excel")
async def export_leave_to_excel(
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    date_range: Optional[str] = Query(None),
    user_ids: Optional[str] = Query(None),  # Comma-separated UUIDs
    db: Session = Depends(get_db),
    _: User = Depends(require_admin)
):
    """Export leave report to Excel (admin only)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl library is not installed"
        )

    # Handle date range
    if not from_date or not to_date:
        if date_range and date_range != "custom":
            from_date, to_date = get_date_range(date_range)
        else:
            from_date, to_date = get_date_range("this_month")

    query = db.query(
        LeaveRequest.from_date,
        LeaveRequest.to_date,
        User.name,
        User.email,
        LeaveRequest.reason,
        LeaveRequest.status,
        LeaveRequest.admin_comment
    ).join(User, LeaveRequest.user_id == User.id).filter(
        LeaveRequest.from_date <= to_date,
        LeaveRequest.to_date >= from_date,
        User.role == 'EMPLOYEE',
    )

    if user_ids:
        user_id_list = [uid.strip() for uid in user_ids.split(',') if uid.strip()]
        if user_id_list:
            query = query.filter(LeaveRequest.user_id.in_(user_id_list))

    results = query.order_by(LeaveRequest.from_date.desc()).all()

    # ── Excel workbook ──────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Leave Report"

    title_fill  = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_font  = Font(color="FFFFFF", bold=True, size=14)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center      = Alignment(horizontal="center", vertical="center")
    border      = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )
    approved_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    rejected_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    pending_fill  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    NUM_COLS = 6
    row_num = 1

    # Title row
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=NUM_COLS)
    tc = ws.cell(row=row_num, column=1, value=f"Leave Report: {from_date} to {to_date}")
    tc.fill, tc.font, tc.alignment = title_fill, title_font, center
    row_num += 2

    # Header row
    headers = ["Employee", "From", "To", "Reason", "Status", "Admin Comment"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=header)
        cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, center, border
    row_num += 1

    if not results:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=NUM_COLS)
        nd = ws.cell(row=row_num, column=1, value="No leave data found for the selected date range and filters")
        nd.alignment = center
        nd.font = Font(italic=True, size=12)
    else:
        for row in results:
            status_val = str(row[5].value if hasattr(row[5], 'value') else row[5])
            row_fill = (
                approved_fill if status_val == 'APPROVED' else
                rejected_fill if status_val == 'REJECTED' else
                pending_fill
            )
            cells = [
                ws.cell(row=row_num, column=1, value=row[2]),
                ws.cell(row=row_num, column=2, value=str(row[0])),
                ws.cell(row=row_num, column=3, value=str(row[1])),
                ws.cell(row=row_num, column=4, value=row[4]),
                ws.cell(row=row_num, column=5, value=status_val),
                ws.cell(row=row_num, column=6, value=row[6] or ''),
            ]
            for cell in cells:
                cell.fill = row_fill
                cell.border = border
            row_num += 1

    # Column widths
    for col, width in zip("ABCDEF", [22, 14, 14, 35, 14, 30]):
        ws.column_dimensions[col].width = width

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    filename = f"leave_report_{from_date}_{to_date}.xlsx"
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# User Reports Router (for non-admin users)
user_reports_router = APIRouter(prefix="/reports", tags=["User - Reports"])


@user_reports_router.get("/my-timesheet", response_model=List[TimesheetReportItem])
def get_my_timesheet_report(
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    date_range: Optional[str] = Query(None),
    client_ids: Optional[str] = Query(None),  # Comma-separated UUIDs
    task_master_ids: Optional[str] = Query(None),  # Comma-separated UUIDs
    is_profitable: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get current user's timesheet report with filters."""
    # Handle date_range filter
    if date_range:
        from_date, to_date = get_date_range(date_range)
    elif not from_date or not to_date:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Either provide from_date and to_date, or date_range"
        )
    
    query = db.query(
        TaskEntry.work_date,
        User.name,
        User.email,
        func.coalesce(Client.name, 'No Client').label('client_name'),
        TaskEntry.task_name,
        TaskEntry.total_hours,
        TaskEntry.status,
        TaskEntry.admin_comment
    ).join(User, TaskEntry.user_id == User.id).outerjoin(
        Client,
        TaskEntry.client_id == Client.id  # LEFT JOIN to include leave entries
    ).filter(
        TaskEntry.work_date >= from_date,
        TaskEntry.work_date <= to_date,
        TaskEntry.user_id == current_user.id  # Only current user's entries
    )
    
    # Multi-select filters
    if client_ids:
        client_id_list = [cid.strip() for cid in client_ids.split(',') if cid.strip()]
        if client_id_list:
            query = query.filter(TaskEntry.client_id.in_(client_id_list))
    
    if task_master_ids:
        task_master_id_list = [tmid.strip() for tmid in task_master_ids.split(',') if tmid.strip()]
        if task_master_id_list:
            query = query.join(TaskSubEntry, TaskEntry.id == TaskSubEntry.task_entry_id)
            query = query.filter(TaskSubEntry.task_master_id.in_(task_master_id_list))
    
    if is_profitable is not None:
        if not task_master_ids:  # Only join if not already joined
            query = query.join(TaskSubEntry, TaskEntry.id == TaskSubEntry.task_entry_id)
        query = query.join(TaskMaster, TaskSubEntry.task_master_id == TaskMaster.id)
        query = query.filter(TaskMaster.is_profitable == is_profitable)
    
    results = query.order_by(TaskEntry.work_date.desc()).all()
    
    return [
        TimesheetReportItem(
            date=row[0],
            employee_name=row[1],
            employee_email=row[2],
            client_name=row[3],
            task_name=row[4],
            total_hours=float(row[5]) if row[5] else 0.0,
            status=row[6],
            admin_comment=row[7] if row[7] else ""
        )
        for row in results
    ]


@user_reports_router.get("/my-export/excel")
async def export_my_report_to_excel(
    from_date: Optional[date] = Query(None),
    to_date: Optional[date] = Query(None),
    date_range: Optional[str] = Query("this_month"),
    client_ids: Optional[str] = Query(None),
    task_master_ids: Optional[str] = Query(None),
    is_profitable: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export current user's productivity report to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl library is not installed"
        )
    
    # Handle date_range filter
    if not from_date or not to_date:
        # If dates not provided, use date_range (but skip "custom")
        if date_range and date_range != "custom":
            from_date, to_date = get_date_range(date_range)
        elif not date_range or date_range == "custom":
            # Default to current month if no parameters provided or date_range is "custom"
            from_date, to_date = get_date_range("this_month")
    
    # Query for user's data (using TaskSubEntry.client_id)
    query = db.query(
        func.coalesce(Client.name, 'No Client').label('client_name'),
        TaskMaster.name.label('task_name'),
        func.sum(TaskSubEntry.production).label('total_production'),
        func.sum(TaskSubEntry.hours).label('total_hours')
    ).join(
        TaskEntry, TaskSubEntry.task_entry_id == TaskEntry.id
    ).outerjoin(
        Client, TaskSubEntry.client_id == Client.id  # Use TaskSubEntry.client_id with LEFT JOIN
    ).join(
        TaskMaster, TaskSubEntry.task_master_id == TaskMaster.id
    ).filter(
        TaskEntry.work_date >= from_date,
        TaskEntry.work_date <= to_date,
        TaskEntry.status == TaskEntryStatus.APPROVED,
        TaskEntry.user_id == current_user.id  # Only current user's data
    )
    
    # Multi-select filters
    if client_ids:
        client_id_list = [cid.strip() for cid in client_ids.split(',') if cid.strip()]
        if client_id_list:
            query = query.filter(TaskSubEntry.client_id.in_(client_id_list))
    
    if task_master_ids:
        task_master_id_list = [tmid.strip() for tmid in task_master_ids.split(',') if tmid.strip()]
        if task_master_id_list:
            query = query.filter(TaskSubEntry.task_master_id.in_(task_master_id_list))
    
    if is_profitable is not None:
        query = query.filter(TaskMaster.is_profitable == is_profitable)
    
    query = query.group_by(Client.name, TaskMaster.name).order_by(Client.name, TaskMaster.name)
    
    results = query.all()
    
    # Group by client
    from collections import defaultdict
    client_data = defaultdict(list)
    for row in results:
        client_data[row.client_name].append({
            'task_name': row.task_name,
            'production': float(row.total_production or 0),
            'hours': float(row.total_hours or 0)
        })
    
    # Create Excel workbook (reuse same styling as admin export)
    wb = Workbook()
    ws = wb.active
    ws.title = "My Productivity Report"
    
    # Styling
    title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_font = Font(color="FFFFFF", bold=True, size=14)
    
    client_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    client_header_font = Font(color="FFFFFF", bold=True, size=12)
    
    column_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    column_header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row_num = 1
    
    # Add title row
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
    title_cell = ws.cell(row=row_num, column=1, value=f"My Productivity Report: {from_date} to {to_date}")
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = header_alignment
    row_num += 2
    
    # Check if there's any data
    if not client_data:
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
        no_data_cell = ws.cell(row=row_num, column=1, value="No data found for the selected filters")
        no_data_cell.alignment = header_alignment
        no_data_cell.font = Font(italic=True, size=12)
        row_num += 2
        
        headers = ["Task Name", "Production", "Hours", "Efficiency"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.fill = column_header_fill
            cell.font = column_header_font
            cell.alignment = header_alignment
            cell.border = border_style
    else:
        # Styling for subtotal rows
        subtotal_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        subtotal_font = Font(bold=True, size=11)
        
        # Styling for grand total
        grand_total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        grand_total_font = Font(bold=True, size=12)
        
        grand_total_production = 0
        
        for client_name, tasks in client_data.items():
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
            client_cell = ws.cell(row=row_num, column=1, value=client_name)
            client_cell.fill = client_header_fill
            client_cell.font = client_header_font
            client_cell.alignment = header_alignment
            row_num += 1
            
            headers = ["Task Name", "Production", "Hours", "Efficiency"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.fill = column_header_fill
                cell.font = column_header_font
                cell.alignment = header_alignment
                cell.border = border_style
            row_num += 1
            
            # Calculate client subtotal
            client_total_production = 0
            client_total_hours = 0
            
            for task in tasks:
                efficiency = (task['production'] / task['hours']) if task['hours'] > 0 else 0
                
                ws.cell(row=row_num, column=1, value=task['task_name']).border = border_style
                ws.cell(row=row_num, column=2, value=task['production']).border = border_style
                ws.cell(row=row_num, column=3, value=task['hours']).border = border_style
                ws.cell(row=row_num, column=4, value=round(efficiency, 2)).border = border_style
                row_num += 1
                
                client_total_production += task['production']
                client_total_hours += task['hours']
            
            # Add Production Sub Total row for this client
            subtotal_cell = ws.cell(row=row_num, column=1, value="Production Sub Total")
            subtotal_cell.fill = subtotal_fill
            subtotal_cell.font = subtotal_font
            subtotal_cell.border = border_style
            
            production_cell = ws.cell(row=row_num, column=2, value=round(client_total_production, 2))
            production_cell.fill = subtotal_fill
            production_cell.font = subtotal_font
            production_cell.border = border_style
            
            hours_cell = ws.cell(row=row_num, column=3, value=round(client_total_hours, 2))
            hours_cell.fill = subtotal_fill
            hours_cell.font = subtotal_font
            hours_cell.border = border_style
            
            client_efficiency = (client_total_production / client_total_hours) if client_total_hours > 0 else 0
            efficiency_cell = ws.cell(row=row_num, column=4, value=round(client_efficiency, 2))
            efficiency_cell.fill = subtotal_fill
            efficiency_cell.font = subtotal_font
            efficiency_cell.border = border_style
            
            row_num += 1
            
            # Add to grand total
            grand_total_production += client_total_production
            
            row_num += 1
        
        # Add Total Production at the end
        row_num += 1  # Extra spacing before grand total
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=1)
        total_label_cell = ws.cell(row=row_num, column=1, value="TOTAL PRODUCTION")
        total_label_cell.fill = grand_total_fill
        total_label_cell.font = grand_total_font
        total_label_cell.alignment = header_alignment
        total_label_cell.border = border_style
        
        total_value_cell = ws.cell(row=row_num, column=2, value=round(grand_total_production, 2))
        total_value_cell.fill = grand_total_fill
        total_value_cell.font = grand_total_font
        total_value_cell.alignment = header_alignment
        total_value_cell.border = border_style
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    filename = f"my_productivity_report_{from_date}_{to_date}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
